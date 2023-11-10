# -*- coding: utf-8 -*-
# ======================================
# 電子マネー管理システム
# MariaデータベースからデータExcel及びPDF出力
# 時間別売上集計
# [環境]20
#   Python 3.10.8
#   VSCode 1.64
#   <拡張>
#     |- Python  V2021.12
#     |- Pylance V2021.12
#
# [更新履歴]
#   2023/11/6  新規作成
# ======================================
from datetime import datetime
import datetime
import jpholiday
import pandas as pd
import openpyxl 
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from emunmerge import excel_operate

####
# 初期処理
####
class dbJikanReport:        
    # クラス初期化
    def __init__(self, df_paylog, file_out_path, flg, sdate, edate, weather_data):  
        
        # クラス初期化
        self.df_paylog = df_paylog 
        self.file_out_path = file_out_path
        
        self.weatherdata = weather_data
        
        self.SYEAR = sdate.year
        self.SMONTH = sdate.month
        self.SDAY = sdate.day
        self.EYEAR = edate.year
        self.EMONTH = edate.month
        self.EDAY = edate.day              
        
        if flg == '1':
            self.sheet_name = '時間別(現金)'
        else:
            self.sheet_name = '時間別(電子決済)'
        self.dfw0 = self.df_paylog[df_paylog['paykbncd'] == flg] 
        if len(self.dfw0) > 0:
            dfw1 = self.dfw0[['paydatestr','paydatedec','payhour','payprice','paytimestr']] 
            dfw2 = dfw1.astype({'paydatedec':int,'paytimestr':str,'payprice': float,'payhour':int}) 
            #日付でソート 
            dfw3 = dfw2.sort_values(by='paydatedec')
            #時間でソート
            dfx = dfw3.sort_values(by='payhour')  
            self.df_paylog1 = pd.pivot_table(dfx, index=['paydatedec'], columns='payhour',values=['payprice'],aggfunc='sum',margins=True)  
            
    ####################
    # 時間別集計表出力
    ####################
    def print_jikan(self):   
        
        #debug
        print('時間別集計表出力開始：',datetime.datetime.now())      
        if len( self.dfw0) <= 0: #データ0件
            return 0
                    
        #データをExcelに出力     
        with pd.ExcelWriter(f'{self.file_out_path}', mode='a') as writer:
            self.df_paylog1.to_excel(writer,startrow=3,startcol=1,sheet_name=self.sheet_name)   
            
        #Excelシートの加工
        wb = openpyxl.load_workbook(f'{self.file_out_path}')
        sh = wb[self.sheet_name]
        
        #天気情報挿入
        sh.insert_cols(3, 4) #列挿入        
        for i in range(7,len(self.weatherdata)+7):
            sh.cell(row=i, column=3).value = self.weatherdata[i-7][1] #天気（昼）
            sh.cell(row=i, column=4).value = self.weatherdata[i-7][2] #天気（夜）
            sh.cell(row=i, column=5).value = self.weatherdata[i-7][3] #最高気温
            sh.cell(row=i, column=6).value = self.weatherdata[i-7][4] #最低気温
                      
        #用紙設定
        wps = sh.page_setup
        # 用紙サイズを設定
        wps.paperSize = sh.PAPERSIZE_A3
        # 印刷の向きを設定wq
        wps.orientation = sh.ORIENTATION_LANDSCAPE
        wps.fitToWidth = 1
        wps.fitToHeight = 1
        sh.sheet_properties.pageSetUpPr.fitToPage = True       
        #ヘッダーセット
        sh.cell(row=1, column=2).value='売上日・時間別売上集計表'
        sh.cell(row=1, column=3).value=self.sheet_name
            
        str1 = (f'{self.SYEAR} 年 {self.SMONTH} 月 {self.SDAY} 日  ～')
        str2 = (f'{self.EYEAR} 年 {self.EMONTH} 月 {self.EDAY} 日')
        sh.cell(row=2, column=2).value=str1
        sh.cell(row=2, column=3).value=str2
                
        # 決済金額合計をセット
        # 最終行の取得
        maxr = sh.max_row
        maxc = sh.max_column
            
        #罫線引く
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
            
        for row_num in range(4,maxr+1):    
            for col_num in range(2,maxc+1):
                sh.cell(row=row_num ,column=col_num).border = border
            
            
        #決済金額のフォーマットを変更
        for i in range(5,maxr+1):
            for j in range(4,maxc+1):
                sh.cell(row=i,column=j).number_format = "#,##0"
            
        sh.cell(row=maxr,column=2).value = "合計"
        sh.cell(row=5,column=maxc).value = "合計"
        sh.cell(row=6,column=2).value = "売上日"
        sh.cell(row=6,column=3).value = "天気（6:00～18:00)"
        sh.cell(row=6,column=4).value = "天気（18:00～翌6:00)"
        sh.cell(row=6,column=5).value = "最高気温"
        sh.cell(row=6,column=6).value = "最低気温"
            
        # セル幅を自動調整
        for col in sh.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))        
                
            adjusted_width = (max_length + 1) * 1.3   
            sh.column_dimensions[col[0].column_letter].width = adjusted_width
                
        #部分的にセル幅を修正
        sh.column_dimensions['B'].width = 25 #売上日
        sh.column_dimensions['C'].width = 25 #天気（昼）
        sh.column_dimensions['D'].width = 25 #天気（夜）
        sh.column_dimensions['E'].width = 10 #最高気温
        sh.column_dimensions['F'].width = 10 #最低気温
        sh.column_dimensions['G'].width = 7
        #表記を修正
        sh.cell(row=5, column=2).value=' '  
        sh.cell(row=5, column=6).value='時間  →' 
                    
        #祝祭日の背景に色をつける
        # 会社特有の休日
        #company_holiday = ['2018-01-02','2018-01-03','2018-12-28','2018-12-31']
        ymdx = 99999999
        for row_num in range(7,maxr):
            ymd = sh.cell(row=row_num,column=2).value
            wmd = str(ymd)
            if ymd != ymdx and ymd != None:
                ymdx = sh.cell(row=row_num,column=2).value   
                y = wmd[0:4]
                m = wmd[4:6]
                d = wmd[6:8]
                sh.cell(row=row_num,column=2).number_format = "###0"
                dt = datetime.date(int(y),int(m),int(d))
                cel = sh.cell(row=row_num,column=2)           

                #土日、祝祭日判定してセルに色をつける   
                # 通常の土日
                if dt.weekday() == 5:
                    sh[cel.coordinate].fill = PatternFill(patternType='solid', fgColor='ffb76e')
                
                if dt.weekday() == 6:
                    sh[cel.coordinate].fill = PatternFill(patternType='solid', fgColor='ff2d3d')

            # 祝日
            dy = int(y)
            dm = int(m)
            dd = int(d)
            #祝日判定
            res_horiday = jpholiday.is_holiday_name(datetime.date(dy,dm,dd))
            if res_horiday != None:
                sh[cel.coordinate].fill = PatternFill(patternType='solid', fgColor='8eef6e')
        
            # #会社の休日
            # if date.strftime("%Y-%m-%d") in company_holiday:
            #     return Tru                
        
        # save xlsx file
        wb.save(f'{self.file_out_path}')
        
        #debug
        print('時間別集計表出力終了：',datetime.datetime.now())      
        
        return 0
    
    ###############################################################
    # ディストラクタ
    ###############################################################
    def __del__(self):
        #print('ディストラクタ呼び出し') 
        pass 