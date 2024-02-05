##########################################
#
#   営業提出資料
#   見積書データ取得
#   
#   指定期間に作成された見積書のタイトルを取得
#   
#   2024_1_25 新規作成
#   python 3.10.3
#
##########################################

import os
from datetime import datetime, date
import openpyxl
import pyexcel as p
import xlwings as xw

wb = openpyxl.load_workbook(r'C:\Users\user\OneDrive\Workplace\2024年営業計画\売上計画案（東京本社）.xlsx')
sh_jyutyu = wb['受注見込']
rowno = sh_jyutyu.max_row + 1

outcount = 0
file_dir = r'Z:\見積書'
input_fromdate = datetime(2024,1,1,0,0,0)
input_todate = datetime(2024,2,29,0,0,0)

print('処理開始')
# 
for file in os.listdir(file_dir):
    base, ext = os.path.splitext(file)
    if ext == '.xlsx' or ext == '.xls':
        file_path = os.path.join(file_dir, file)      
        file_info = os.stat(file_path)
        create_date = datetime.fromtimestamp(file_info.st_ctime)
        update_date = datetime.fromtimestamp(file_info.st_mtime)
        if create_date >= input_fromdate:
            if create_date <= input_todate:
                sh_jyutyu.cell(rowno,1).value = base
                sh_jyutyu.cell(rowno,2).font = openpyxl.styles.fonts.Font(color='00FFFF')
                sh_jyutyu.cell(rowno,2).hyperlink = file_path
                sh_jyutyu.cell(rowno,3).value = date(int(create_date.year), int(create_date.month), int(create_date.day))
                sh_jyutyu.cell(rowno,4).value = date(int(update_date.year), int(update_date.month), int(update_date.day))
                #rowno += 1
                #outcount += 1
                #########################################################
                #
                # 見積りファイルから金額等を取り出す
                #
                #########################################################
                if ext == '.xls': #.xlsファイルのファイル名を一時的に変換
                    wk_filepath = base + '.xlsx'
                    new_file = os.path.join(file_dir,wk_filepath)
                    #p.save_book_as(file_name=file_path, dest_file_name=new_file)
                    #オブジェクト作成し、新規ブック作成＝Excel自動起動
                    wb = xw.Book()
                    wb = xw.Book(file_path) #現在ファイルの読込み
                    wb.save(new_file)
                    wb.close()
                    flg = 1
                else:
                    new_file = file_path
                    flg = 0

                new_base, new_ext = os.path.splitext(new_file)

                if new_ext == '.xlsx':
                    try:
                        wbm = openpyxl.load_workbook(new_file,data_only=True)
                        wk_companyname = ""
                        wk_sales = 0
                        for shm in wbm.sheetnames:
                            print('シートの読込 =',shm)
                            if shm == 'Sheet1': #sheet1のみ検索対象とする    
                                sh_mitu = wbm[shm]                   
                                for row_no in range(1,27):
                                    for col_no in range(1,12):
                                        wk_str = sh_mitu.cell(row_no,col_no).value
                                        #if shm.cell(row_no,col_no).value == '御中': #社名検索
                                        if wk_str == '御中': #社名検索
                                            wk_companyname = sh_mitu.cell(row_no,col_no-4).value
                                            print('社名  :',wk_companyname)
                                        if sh_mitu.cell(row_no,col_no).value == '総金額': #総額検索
                                            if sh_mitu.cell(row_no,col_no+2).value != None:
                                                if sh_mitu.cell(row_no,col_no+2).value >= 0:
                                                    wk_sales = int(sh_mitu.cell(row_no,col_no+2).value)
                                                    print('金額  :',wk_sales)
                    except FileNotFoundError:
                        print('ファイルが読み込めない New_File = ',new_file)
                        if flg == 1:
                            print('変換前ファイル名 = ',file_path)
                    if wk_companyname != "":
                        sh_jyutyu.cell(rowno,6).value = wk_companyname
                    if wk_sales >= 0:
                        sh_jyutyu.cell(rowno,7).value = wk_sales

                    if flg == 1: #一次的に変換したファイルを削除                 
                        os.remove(new_file)

                rowno += 1
                outcount += 1

print('出力件数', outcount)
wb.save(r'C:\Users\user\OneDrive\Workplace\2024年営業計画\売上計画案（東京本社）.xlsx')