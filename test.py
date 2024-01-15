from emoneydbclass import DataBaseClass
import yaml
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

print('処理開始：',datetime.datetime.now()) 

parm_data = []    
# 基本情報取得
with open('C:/emoney/emoney.yaml','r+',encoding="utf-8") as ry:
    config_yaml = yaml.safe_load(ry)
    dbip = config_yaml['dbip']
    dbmarianame = config_yaml['dbmarianame']
    dbport = config_yaml['dbport']        
    dbuser = config_yaml['dbuser']
    dbpw = config_yaml['dbpw']
    parm_data.append(dbip)
    parm_data.append(dbmarianame)
    parm_data.append(dbport)        
    parm_data.append(dbuser)
    parm_data.append(dbpw)    
    file_path = config_yaml['dir_filepath']
    parm_data.append(file_path)
    
    #データベース操作クラス初期化
    resdb = DataBaseClass(parm_data)
    company = '0000004'
    start_date = datetime.datetime(2022,5,1)
    end_date = datetime.datetime(2023,10,31)
    #データ切出し
    ret_paylog = resdb.paylog_sum_get(company,start_date,end_date)
    
    #データをExcelに出力    
    sheet_name = 'test'
    file_out_path = 'c:/emoney/test.xlsx'
    with pd.ExcelWriter(f'{file_out_path}', mode='a') as writer:
        ret_paylog.to_excel(writer,startrow=3,startcol=1,sheet_name=sheet_name) 
    wb = openpyxl.load_workbook(f'{file_out_path}')
    sh = wb[sheet_name]
    
    #用紙設定
    wps = sh.page_setup
    # 用紙サイズを設定
    wps.paperSize = sh.PAPERSIZE_A3
    # 印刷の向きを設定
    wps.orientation = sh.ORIENTATION_LANDSCAPE
        
    sh.cell(row=1, column=2).value='月別設置場所別集計表'
    sh.cell(row=1, column=3).value=sheet_name
        
    str1 = (f'{start_date.year} 年 {start_date.month} 月 {start_date.day} 日  ～')
    str2 = (f'{end_date.year} 年 {end_date.month} 月 {end_date.day} 日')
    sh.cell(row=2, column=2).value=str1
    sh.cell(row=2, column=3).value=str2
    
    # 最終行・列数の取得
    maxr = sh.max_row
    maxc = sh.max_column
    #表記を修正
    sh.cell(row=6, column=2).value='決済年'
    sh.cell(row=6, column=3).value='決済月'  
    sh.cell(row=5, column=3).value=''  
    sh.cell(row=4, column=4).value='設置場所' 
    sh.cell(row=maxr,column=2).value = '合計'
    sh.cell(row=5,column=maxc).value = '合計'
    
    #金額の表示フォーマットを変更
    for i in range(4,maxr+1):
        for j in range(4,maxc+1):
            sh.cell(row=i,column=j).number_format = "#,##0"
    
    #タイトルの文字サイズ変更
    font = Font(name='Yu Gothic', sz = 8)
    for j in range(4,maxc):        
        sh.cell(row=5,column=j).font = font
    
    # セル幅を自動調整
    for col in sh.columns:
        max_length = 0
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))  
            adjusted_width = (max_length + 1) * 1.3  
            sh.column_dimensions[col[0].column_letter].width = adjusted_width
    
    #部分的にセル幅を修正
    sh.column_dimensions['B'].width = 20
    sh.column_dimensions['C'].width = 15     
    
    #罫線引く
    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
        
    for row_num in range(6,maxr+1):    
        for col_num in range(4,maxc+1):
            sh.cell(row=row_num ,column=col_num).border = border

    wb.save(f'{file_out_path}')
    
    del resdb 