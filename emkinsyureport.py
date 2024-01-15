# -*- coding: utf-8 -*-
# ======================================
# 電子マネー管理システム
# MariaデータベースからデータExcel及びPDF出力
# 金種別売上集計
# [環境]20
#   Python 3.10.8
#   VSCode 1.64
#   <拡張>
#     |- Python  V2021.12
#     |- Pylance V2021.12
#
# [更新履歴]
#   2023/11/6  新規作成
# ======================================
from datetime import datetime
import datetime
import jpholiday
import pandas as pd
import openpyxl 
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from emunmerge import excel_operate

####
# 初期処理
####
class dbKinsyuReport:        
    # クラス初期化
    def __init__(self, df_paylog, file_out_path, flg, sdate, edate):  
        
        # クラス初期化
        self.df_paylog = df_paylog 
        self.file_out_path = file_out_path
        
        self.SYEAR = sdate.year
        self.SMONTH = sdate.month
        self.SDAY = sdate.day
        self.EYEAR = edate.year
        self.EMONTH = edate.month
        self.EDAY = edate.day              
        
        if flg == '1':
            self.sheet_name = '金種別(現金)'
        else:
            self.sheet_name = '金種別(電子決済)'
        self.dfw0 = df_paylog[df_paylog['paykbncd'] == flg] 
        if len(self.dfw0) > 0:              
            dfw1 = self.dfw0[['paydatedec','payhour','payprice','paytimestr']]
            dfw2 = dfw1.astype({'paytimestr':str,'payprice': float,'payhour':int})  
            # 日付でソート
            dfw3 = dfw2.sort_values(by='paydatedec')
            # 時間でソート
            dfx = dfw3.sort_values(by='payhour') 
            #日付・時間で決済金額を集計                
            self.df_paylog1 = pd.pivot_table(dfx, index=['paydatedec','payprice'], columns='payhour',aggfunc='count',margins=True,margins_name='Total')        
            #金種・日付で決済件数を集計
            self.df_paylog2 = pd.pivot_table(dfx, index=['payprice'], columns='paydatedec',aggfunc='count',margins=True,margins_name='Total')
            
     ##########################
    # 金種別選択回数集計表出力
    # ########################  
    def print_kinsyu(self):
        
        #debug
        print('金種別集計表出力開始：',datetime.datetime.now()) 
        if len(self.dfw0) <= 0: #データ0件
            return 0
        
        #データをExcelに出力
        sheet_name2 = self.sheet_name + '_2'
        with pd.ExcelWriter(f'{self.file_out_path}', mode='a') as writer:
            self.df_paylog1.to_excel(writer,startrow=3,startcol=1,sheet_name=self.sheet_name)  
            self.df_paylog2.to_excel(writer,startrow=1,startcol=1,sheet_name=sheet_name2) 

        # #日付欄の結合を外す
        Inputfile = f'{self.file_out_path}'
        Excel = excel_operate(Inputfile)
        Excel.get_merged_cells_location()
        Excel.break_merged_cells() 

        #Excelシートの加工
        wb = openpyxl.load_workbook(f'{self.file_out_path}')
        sh = wb[self.sheet_name]
        shx = wb[sheet_name2]
        
        #表記を修正
        sh.cell(row=4, column=4).value='決済時刻'
        sh.cell(row=5, column=3).value=''  
        #sh.cell(row=6, column=3).value='決済種別'  
        sh.cell(row=6, column=3).value='決済金額'      
        
        # 最終行の取得
        maxr = sh.max_row
        maxc = sh.max_column
        maxr2 = shx.max_row+1
        #一部結合中央揃え    
        sh.merge_cells(start_row=4,start_column=4,end_row=4,end_column=maxc) 
                    
        #用紙設定
        wps = sh.page_setup
        # 用紙サイズを設定
        wps.paperSize = sh.PAPERSIZE_A3
        # 印刷の向きを設定
        wps.orientation = sh.ORIENTATION_LANDSCAPE
            
        sh.cell(row=1, column=2).value='金種別利用回数集計表'
        sh.cell(row=1, column=3).value=self.sheet_name
            
        str1 = (f'{self.SYEAR} 年 {self.SMONTH} 月 {self.SDAY} 日  ～')
        str2 = (f'{self.EYEAR} 年 {self.EMONTH} 月 {self.EDAY} 日')
        sh.cell(row=2, column=2).value=str1
        sh.cell(row=2, column=3).value=str2
            
        # セル幅を自動調整
        for col in sh.columns:
            max_length = 0
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))        
                
            adjusted_width = (max_length + 1) * 1.3   
            sh.column_dimensions[col[0].column_letter].width = adjusted_width
                
        #部分的にセル幅を修正
        sh.column_dimensions['B'].width = 25 #売上日
        #sh.column_dimensions['C'].width = 15 #決済種別
        sh.column_dimensions['C'].width = 15 #決済金額    
        #罫線引く
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
            
        for row_num in range(4,maxr+1):    
            for col_num in range(3,maxc+1):
                sh.cell(row=row_num ,column=col_num).border = border
            
            
        #決済金額の表示フォーマットを変更
        for i in range(7,maxr+1):
            for j in range(2,maxc+1):
                #金種
                sh.cell(row=i,column=j).number_format = "#,##0"
        for i in range(7,maxr+1):
            sh.cell(row=i,column=3).number_format = "#,##0"

        sh.cell(row=maxr,column=2).value = "合計"
        sh.cell(row=5,column=maxc).value = "合計"
        sh.cell(row=6,column=2).value = "決済日"
            
        #金種別の合計表示
        ft = Font(bold=True)
        sh.cell(row=maxr+2,column=2).value = "金種別件数合計"
        sh.cell(row=maxr+2,column=2).font = ft
        #for i in range(5,12):
        for i in range(5,maxr2):
        #金種
            copy = shx.cell(row = i, column =2).value
            sh.cell(row = maxr+i-2, column = 3 , value = copy) .number_format = "#,##0"
            sh.cell(row = maxr+i-2, column = 3 , value = copy).font = ft
            sh.cell(row = maxr+i-2, column = 3 , value = copy).alignment = Alignment(horizontal = 'center', 
                                                                                            vertical = 'center')
        #件数合計
        for v in range(3,41):
            if shx.cell(row = 3, column = v).value == "Total": #合計行検索
                for p in range(4,maxr2):
                    copy2 = shx.cell(row = p, column = v).value
                    sh.cell(row = maxr+p-2, column = maxc , value = copy2).number_format = "#,##0"
                break
                
        #祝祭日の背景に色をつける
        # 会社特有の休日
        #company_holiday = ['2018-01-02','2018-01-03','2018-12-28','2018-12-31']        
        ymdx = 99999999
        for row_num in range(7,maxr):
            ymd = sh.cell(row=row_num,column=2).value
            wmd = str(ymd)
            if ymd != ymdx and ymd != None:
                #sh.unmerge_cells(row=row_num,column=2)
                ymdx = sh.cell(row=row_num,column=2).value   
                y = wmd[0:4]
                m = wmd[4:6]
                d = wmd[6:8]
                sh.cell(row=row_num,column=2).number_format = "###0"
                dt = datetime.date(int(y),int(m),int(d))
                cel = sh.cell(row=row_num,column=2)

                #土日、祝祭日判定してセルに色をつける   
                # 通常の土日
                if dt.weekday() == 5:
                    sh[cel.coordinate].fill = PatternFill(patternType='solid', fgColor='ffb76e')
            
                if dt.weekday() == 6:
                    sh[cel.coordinate].fill = PatternFill(patternType='solid', fgColor='ff2d3d')

                # 祝日
                dy = int(y)
                dm = int(m)
                dd = int(d)
                #祝日判定
                res_horiday = jpholiday.is_holiday_name(datetime.date(dy,dm,dd))
                if res_horiday != None:
                    sh[cel.coordinate].fill = PatternFill(patternType='solid', fgColor='8eef6e')
                    
        #ワーク用シートの削除とブックの保存
        #wb.remove(wb.worksheets[3])
        wb.remove(wb[sheet_name2])
        wb.save(f'{self.file_out_path}')
        
        #debug
        print('金種別集計表出力終了：',datetime.datetime.now()) 
        
        return 0 
    
    ###############################################################
    # ディストラクタ
    ###############################################################
    def __del__(self):
        #print('ディストラクタ呼び出し') 
        pass 