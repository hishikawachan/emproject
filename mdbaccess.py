#################################################
#
#   営業提出資料
#   納品書データ取得
#   保守データ取得
#
#   指定期間に作成された納品書を取得し売上実績を把握
#
#   2024_1_24 新規作成
#   python 3.10.3
#
#################################################

import pyodbc as pyo
import openpyxl
from datetime import date

##########################################
#
#  納品書DB　アクセス
#
##########################################

con_str1 = (
	r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};'
	r'DBQ=Z:\データベース\納品書\納品書(税率10％ 西暦表示 版)\納品書Ver.1.2元データ用(2024版 税率10% 西暦表示).mdb;'
	)

con = pyo.connect(con_str1)
cursor = con.cursor()
###################################
# 納品書データを日付で抽出
# 指定日付はSQL文を変更
###################################
sql1 = 'SELECT * FROM 納品書 \
        LEFT JOIN ゴルフ場名簿 ON(納品書.ゴルフ場No = ゴルフ場名簿.ゴルフ場No) \
        WHERE 納品日 Between #2024/01/01# AND #2024/01/31#'
rows_nouhin = cursor.execute(sql1).fetchall()

#print(f'納品No={row.納品No}, 納品合計金額={row.納品合計金額}, 摘要={row.摘要}, 納品日={row.納品日.strftime("%Y/%m/%d")}')

rows_nouhin_len = len(rows_nouhin)
###################################
#  売上実績を計画表に書き込み
###################################
wb = openpyxl.load_workbook(r'C:\Users\user\OneDrive\Workplace\2024年営業計画\売上計画案（東京本社）.xlsx')
sh_nouhin = wb['売上実績']
rowno = sh_nouhin.max_row + 1
start_rowno = rowno
#print(f'最終行 = {maxr}')
data_no = 0
for data_no in range(0, rows_nouhin_len):
    if rows_nouhin[data_no].納品合計金額 != None:
        if rows_nouhin[data_no].納品合計金額 != 0:
            sh_nouhin.cell(rowno,1).value = rows_nouhin[data_no].納品No
            #sh_nouhin.cell(rowno,2).value = rows_nouhin[data_no].納品日
            w_date = date(int(rows_nouhin[data_no].納品日.year), int(rows_nouhin[data_no].納品日.month), int(rows_nouhin[data_no].納品日.day))
            sh_nouhin.cell(rowno,2).value = w_date
            sh_nouhin.cell(rowno,3).value = rows_nouhin[data_no].ゴルフ場名
            sh_nouhin.cell(rowno,4).value = rows_nouhin[data_no].摘要
            sh_nouhin.cell(rowno,5).value = rows_nouhin[data_no].納品合計金額 
            sh_nouhin.cell(rowno,6).value = w_date.year
            sh_nouhin.cell(rowno,7).value = w_date.month
            rowno += 1
            
wb.save(r'C:\Users\user\OneDrive\Workplace\2024年営業計画\売上計画案（東京本社）.xlsx')

cursor.close()
con.close()

###################################
# 保守伝票データを日付で抽出
###################################
con_str2 = (
	r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};'
	r'DBQ=Z:\データベース\修理伝票\元データリンク版\修理伝票データベースVr.1.3元データリンク版 - 2023 - .mdb;'
	)
###################################
# 指定日付はSQL文を変更
###################################
sql2 = 'SELECT * FROM ( 修理表 \
        LEFT JOIN 受付表 \
        ON(修理表.受付No = 受付表.受付No)) \
        LEFT JOIN ゴルフ場名簿 \
        ON(受付表.ゴルフ場No = ゴルフ場名簿.ゴルフ場No) \
        WHERE 発行日 Between #2023/12/01# AND #2024/01/31#'

con = pyo.connect(con_str2)
cursor = con.cursor()

rows_hosyu = cursor.execute(sql2).fetchall()

rows_hosyu_len = len(rows_hosyu)

wb = openpyxl.load_workbook(r'C:\Users\user\OneDrive\Workplace\2024年営業計画\売上計画案（東京本社）.xlsx')
sh_hosyu = wb['保守実績']
rowno = sh_hosyu.max_row + 1
start_rowno = rowno

data_no = 0
for data_no in range(0, rows_hosyu_len):
    if rows_hosyu[data_no].総合計 != None:
        if rows_hosyu[data_no].総合計 != 0:
            wh_date = date(int(rows_hosyu[data_no].発行日.year), int(rows_hosyu[data_no].発行日.month), int(rows_hosyu[data_no].発行日.day))
            sh_hosyu.cell(rowno,1).value = wh_date
            ws_date = date(int(rows_hosyu[data_no].修理日.year), int(rows_hosyu[data_no].修理日.month), int(rows_hosyu[data_no].修理日.day))
            sh_hosyu.cell(rowno,2).value = ws_date
            sh_hosyu.cell(rowno,3).value = rows_hosyu[data_no].ゴルフ場名
            sh_hosyu.cell(rowno,4).value = rows_hosyu[data_no].作業内容
            sh_hosyu.cell(rowno,5).value = rows_hosyu[data_no].総合計 
            sh_hosyu.cell(rowno,6).value = wh_date.year
            sh_hosyu.cell(rowno,7).value = wh_date.month
            sh_hosyu.cell(rowno,8).value = 20
            sh_hosyu.cell(rowno,9).value = 400
            rowno += 1

wb.save(r'C:\Users\user\OneDrive\Workplace\2024年営業計画\売上計画案（東京本社）.xlsx')
cursor.close()
con.close()