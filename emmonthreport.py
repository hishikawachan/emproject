# -*- coding: utf-8 -*-
# ======================================
# 電子マネー管理システム
# MariaデータベースからデータExcel及びPDF出力
# 月別売上集計
# [環境]20
#   Python 3.10.8
#   VSCode 1.64
#   <拡張>
#     |- Python  V2021.12
#     |- Pylance V2021.12
#
# [更新履歴]
#   2023/11/6  新規作成
# ======================================
from datetime import datetime
import datetime
import pandas as pd
import openpyxl 
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from emunmerge import excel_operate

####
# 初期処理
####
class dbMonthReport:        
    # クラス初期化
    def __init__(self, df_syubetu, df_paylog, file_out_path, flg, sdate, edate):  
        
        # クラス初期化
        self.df_card = df_syubetu
        self.df_paylog = df_paylog 
        self.file_out_path = file_out_path
        
        self.SYEAR = sdate.year
        self.SMONTH = sdate.month
        self.SDAY = sdate.day
        self.EYEAR = edate.year
        self.EMONTH = edate.month
        self.EDAY = edate.day             
        
        # if flg == '1':
        #     self.sheet_name = '月別(現金)'
        # else:
        #     self.sheet_name = '月別(電子決済)'
        
        self.sheet_name = '月別決済種別'
        # 決済種別を結合
        self.df_paylog = pd.merge(self.df_paylog,self.df_card, left_on='paycardcd', right_on='cardcode') 
        
        #self.df_paylog.sort_index(axis=1, ascending=False, inplace=True)
        self.df_paylog.sort_values(['payyear', 'paymonth', 'cardcode'])
        
        
        #self.dfw0 = df_paylog[df_paylog['paykbncd'] == flg] 
        
        #self.dfw0['pricename'] = str(self.dfw0['payprice'])
        
        if len(self.df_paylog) > 0:
            #self.df_paylog = pd.pivot_table(self.dfw0, index=['payyear','paymonth'], columns=['placename'],values=['payprice'],aggfunc='sum',margins=True,margins_name='Total')  
            self.df_paylog = pd.pivot_table(self.df_paylog, index=['payyear','paymonth'], columns=['cardname'],values=['payprice'],aggfunc='sum',margins=True,margins_name='Total')
 
    ####################
    # 月別決済種別集計表出力
    ####################
    def print_monthly(self): 
        #debug
        print('月別決済種別集計表出力開始：',datetime.datetime.now())  
        if len(self.df_paylog) <= 0: #データ0件
            return 0        
            
        with pd.ExcelWriter(f'{self.file_out_path}', mode='a') as writer:
            self.df_paylog.to_excel(writer,startrow=3,startcol=1,sheet_name=self.sheet_name) 
        wb = openpyxl.load_workbook(f'{self.file_out_path}')
        sh = wb[self.sheet_name]
        
        #用紙設定
        wps = sh.page_setup
        # 用紙サイズを設定
        wps.paperSize = sh.PAPERSIZE_A3
        # 印刷の向きを設定
        wps.orientation = sh.ORIENTATION_LANDSCAPE
            
        sh.cell(row=1, column=2).value='月別決済種別集計表'
        sh.cell(row=1, column=3).value=self.sheet_name
            
        str1 = (f'{self.SYEAR} 年 {self.SMONTH} 月 {self.SDAY} 日  ～')
        str2 = (f'{self.EYEAR} 年 {self.EMONTH} 月 {self.EDAY} 日')
        sh.cell(row=2, column=2).value=str1
        sh.cell(row=2, column=3).value=str2
        
        # 最終行・列数の取得
        maxr = sh.max_row
        maxc = sh.max_column
        #表記を修正
        sh.cell(row=6, column=2).value='決済年'
        sh.cell(row=6, column=3).value='決済月'  
        sh.cell(row=5, column=3).value=''  
        sh.cell(row=4, column=4).value='決済種別' 
        sh.cell(row=maxr,column=2).value = '合計'
        sh.cell(row=5,column=maxc).value = '月合計'
        
        #金額の表示フォーマットを変更
        for i in range(4,maxr+1):
            for j in range(4,maxc+1):
                sh.cell(row=i,column=j).number_format = "#,##0"
        
        #タイトルの文字サイズ変更
        font = Font(name='Yu Gothic', sz = 8)
        for j in range(4,maxc):        
            sh.cell(row=5,column=j).font = font
        
        # セル幅を自動調整
        for col in sh.columns:
            max_length = 0
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))  
                adjusted_width = (max_length + 1) * 1.5 
                sh.column_dimensions[col[0].column_letter].width = adjusted_width
        
        #部分的にセル幅を修正
        sh.column_dimensions['B'].width = 24
        sh.column_dimensions['C'].width = 15     
        
        #罫線引く
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
            
        for row_num in range(6,maxr+1):    
            for col_num in range(4,maxc+1):
                sh.cell(row=row_num ,column=col_num).border = border

        wb.save(f'{self.file_out_path}')
        
        #debug
        print('月別決済種別集計表出力終了：',datetime.datetime.now())   
        
        return 0
    
    ###############################################################
    # ディストラクタ
    ###############################################################
    def __del__(self):
        #print('ディストラクタ呼び出し') 
        pass 