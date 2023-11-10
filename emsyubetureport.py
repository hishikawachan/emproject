# -*- coding: utf-8 -*-
# ======================================
# 電子マネー管理システム
# MariaデータベースからデータExcel及びPDF出力
# 決済種別別売上集計表出力
# [環境]20
#   Python 3.10.8
#   VSCode 1.64
#   <拡張>
#     |- Python  V2021.12
#     |- Pylance V2021.12
#
# [更新履歴]
#   2023/11/6  新規作成
# ======================================
from datetime import datetime
import datetime
import pandas as pd
import openpyxl 
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from emunmerge import excel_operate

####
# 初期処理
####
class dbSyubetuReport:
    def __init__(self,df_syubetu, df_paylog, file_out_path, sdate, edate):  
        
        # クラス初期化
        self.df_card = df_syubetu
        self.df_paylog = df_paylog 
        self.file_out_path = file_out_path 
        
        self.SYEAR = sdate.year
        self.SMONTH = sdate.month
        self.SDAY = sdate.day
        self.EYEAR = edate.year
        self.EMONTH = edate.month
        self.EDAY = edate.day                     
        
    ########################
    # カード種類別集計表
    ########################
    def print_syubetsu(self):
        #debug
        print('カード種類別集計表出力開始：',datetime.datetime.now()) 
        if len(self.df_paylog) <= 0: #データ0件
            return 0
        
        # 決済種別を結合
        df_paylog = pd.merge(self.df_paylog,self.df_card, left_on='paycardcd', right_on='cardcode') 
        # 決済種別毎に集計
        df_sum_card = df_paylog[['cardname','payprice']].groupby('cardname').sum()
        
        #データをExcelに出力    
        with pd.ExcelWriter(f'{self.file_out_path}') as writer:
            df_sum_card.to_excel(writer,startrow=3,startcol=1,sheet_name='決済種別')  

        #Excelシートの加工
        wb = openpyxl.load_workbook(f'{self.file_out_path}')
        sh = wb.worksheets[0]

        #用紙設定
        wps = sh.page_setup
        # 用紙サイズを設定
        wps.paperSize = sh.PAPERSIZE_A4
        # 印刷の向きを設定
        wps.orientation = sh.ORIENTATION_PORTRAIT

        sh.cell(row=1, column=2).value='決済種別売上集計表'
        #sh.cell(row=1, column=3).value='（電子マネー）'

        str1 = (f'{self.SYEAR} 年 {self.SMONTH} 月 {self.SDAY} 日  ～')
        str2 = (f'{self.EYEAR} 年 {self.EMONTH} 月 {self.EDAY} 日')
        sh.cell(row=2, column=2).value=str1
        sh.cell(row=2, column=3).value=str2

        # セル幅を自動調整
        for col in sh.columns:
            max_length = 0
            column = col[0].column

            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))        
        
        adjusted_width = (max_length + 1) * 1.0   
        sh.column_dimensions[col[0].column_letter].width = adjusted_width
        
        #表記を修正
        sh.cell(row=4, column=2).value='決済種別'
        sh.cell(row=4, column=3).value='決済金額'    

        #部分的にセル幅を修正
        sh.column_dimensions['B'].width = 30 #明細種別名称
        sh.column_dimensions['C'].width = 30 #決済金額

        #決済金額合計をセット
        # 最終行の取得
        maxr = sh.max_row

        # 計算用の変数goukeiを定義
        goukei = 0

        # 所定範囲の数値を計算
        for i in range(5, maxr+1):
            kingaku = sh['C' + str(i)].value
            goukei += int(kingaku)

        # 合計値をセルに出力 中央ぞろえ
        sh['C' + str(maxr+1)].value = goukei
        sh['B' + str(maxr+1)].value = "合  計"
        font = Font(bold=True)
        sh['B' + str(maxr+1)].font = font
        sh['B' + str(maxr+1)].alignment = Alignment(horizontal="centerContinuous")

        #罫線引く
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)

        for row_num in range(4,maxr+2):
            for col_num in range(2,4):
                sh.cell(row=row_num ,column=col_num).border = border

        #決済金額のフォーマットを変更
        for i in range(5,maxr+2):
            sh.cell(row=i,column=3).number_format = "¥#,##0" 
        
        wb.save(f'{self.file_out_path}')
        
        #debug
        print('カード種類別集計表出力終了：',datetime.datetime.now()) 
        
        return 0
    
    ###############################################################
    # ディストラクタ
    ###############################################################
    def __del__(self):
        #print('ディストラクタ呼び出し') 
        pass 
        
    