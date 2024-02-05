import os 
import openpyxl
import pyexcel as p
import docx

file_dir = r'c:\test'
file_name = 't e s t.txt'
file_name2 = 'テスト　　テスト　　テスト.xlsx'
file_name3 = 'test.docx'
file_name4 = 'test.xlsx'
file_name5 = 'テスト　test テスト test　.xlsx'

file_path1 = os.path.join(file_dir, file_name)
file_path2 = os.path.join(file_dir, file_name2) 
file_path3 = os.path.join(file_dir, file_name3) 
file_path4 = os.path.join(file_dir, file_name4) 
file_path5 = os.path.join(file_dir, file_name5) 

#file_path2.replace('　', ' ')

print('file_path1 =', file_path1)
print('file_path2 =', file_path2)
print('file_path3 =', file_path3)
print('file_path4 =', file_path4)
print('file_path5 =', file_path5)

document = docx.Document(file_path3)
f = open(file_path1)
wb2 = openpyxl.load_workbook(file_path2,data_only=True)
wb4 = openpyxl.load_workbook(file_path4,data_only=True)
wb5 = openpyxl.load_workbook(file_path5,data_only=True)

print('テスト完了')
